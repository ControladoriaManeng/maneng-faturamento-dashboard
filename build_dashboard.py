# -*- coding: utf-8 -*-
"""
Baixa a planilha oficial de faturamento (Google Drive, link publico),
agrega as notas por SIGLA/mes/tipo de servico (P/C/E) e gera o HTML
final do dashboard a partir do template.

Rodar de novo a qualquer momento reflete o estado atual da planilha.
"""
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone, timedelta

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_URL = "https://docs.google.com/spreadsheets/d/12w6lJOdcIggtQpymebiHp5nbJijpdOQQ/export?format=xlsx"
XLSX_PATH = os.path.join(BASE_DIR, "_faturamento_cache.xlsx")
OSP_URL = "https://docs.google.com/spreadsheets/d/1qo36gG7jvtgO_Z0wgdIGq9ISE3-CKwLd/export?format=xlsx"
OSP_PATH = os.path.join(BASE_DIR, "_osp_cache.xlsx")
TEMPLATE_PATH = os.path.join(BASE_DIR, "template.html")
OUTPUT_PATH = os.path.join(BASE_DIR, "index.html")

MESES_ORDEM = ['JAN-26', 'FEV-26', 'MAR-26', 'ABR-26', 'MAI-26', 'JUN-26',
               'JUL-26', 'AGO-26', 'SET-26', 'OUT-26', 'NOV-26', 'DEZ-26']

# aliases conhecidos de SIGLA (variam entre meses na planilha)
ALIAS = {
    "SAM'S": "SAMS", "GBRABOSA": "GBARBOSA", "BM": "BANCO MERCANTIL", "BCO MERC": "BANCO MERCANTIL",
    "BCO MERCANTIL": "BANCO MERCANTIL", "CARREFOUR": "CRF", "BOTICARIO": "BOT",
    "O BOTICARIO": "BOT",
    "BRADESCO": "BRAD", "NUC BRAD": "IN HAUS",
    "BRADESCO IN HAUS": "IN HAUS",
    "SHOPPING PAMPLONA": "CRF", "SHOP PAMPLONA": "CRF", "MERCANTIL MERCADO": "MERCANTIL",
    "ATAKARE": "ATAKAREJO", "BAUDUCCO2": "BAUDUCCO",
    # descobertos lendo a planilha real em 2026-08-11:
    "BRAD AG": "AG BRAD", "BRADESCO AGENCIAS": "AG BRAD",
    "FUND BRADESCO": "FUND BRAD", "FUND. BRAD": "FUND BRAD",
    "BRADESCO FUNDAÇÃO": "FUND BRAD", "FUNDAÇÃO BRADESCO": "FUND BRAD",
    "BK": "ZAMP",
    "LEROY": "LM",
    "LEROY MERLIN RESIDENTE": "LM_RESIDENTES",
    "SMART FIT": "SMTF",
    # projetos avulsos de Engenharia Bradesco lançados com sigla descritiva em vez de código:
    "BRADESCO - CACOAL": "BRAD", "BRADESCO - PRÉDIO PRATA": "BRAD",
    "BRADESCO - PRÉDIO PRATA SALA LAN": "BRAD",
    "BRADESCO RESERVA TÉCNICA": "BRAD", "BRADESCO RESERVA TÉCNICA - PREVENTIVA": "BRAD",
    "BRADESCO FUNDAÇÃO - PROPRIÁ": "FUND BRAD",
}

# clientes que a Maneng decidiu parar de acompanhar no painel (contrato encerrado etc.) —
# ficam fora mesmo que tenham faturamento historico em meses anteriores.
SIGLAS_IGNORADAS = {
    'ATENTO',
}

# SIGLA -> {nome, obj_key (chave em OBJ_CLIENTE, ou None se sem meta), grupo, icon, cor, bg}
SIGLA_MAP = {
    'AG BRAD': {'nome': 'Agências Bradesco', 'obj_key': None, 'grupo': 'AG BRAD', 'icon': 'B', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'BOT': {'nome': 'O Boticário', 'obj_key': 'O BOTICARIO', 'grupo': 'BOT', 'icon': 'O', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'ZAMP': {'nome': 'ZAMP', 'obj_key': None, 'grupo': 'ZAMP', 'icon': 'Z', 'cor': '#C0200E', 'bg': '#FEF0EE'},
    'CRF': {'nome': 'Carrefour', 'obj_key': 'CARREFOUR', 'grupo': 'normal', 'icon': 'C', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'FUND BRAD': {'nome': 'Bradesco Fundação', 'obj_key': 'BRADESCO FUNDAÇÃO', 'grupo': 'normal', 'icon': 'BF', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'IN HAUS': {'nome': 'Bradesco In Haus', 'obj_key': 'BRADESCO IN HAUS', 'grupo': 'normal', 'icon': 'BI', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'BRAD': {'nome': 'Bradesco Eng.', 'obj_key': None, 'grupo': 'normal', 'icon': 'BE', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'LM': {'nome': 'Leroy Merlin', 'obj_key': 'LEROY', 'grupo': 'normal', 'icon': 'L', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'LM_RESIDENTES': {'nome': 'Leroy Merlin (Residentes)', 'obj_key': 'LEROY MERLIN RESIDENTES', 'grupo': 'normal', 'icon': 'LR', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'SMTF': {'nome': 'Smart Fit', 'obj_key': 'SMART FIT', 'grupo': 'normal', 'icon': 'SF', 'cor': '#E84B1A', 'bg': '#FEF3EE'},
    'GBARBOSA': {'nome': 'GBarbosa', 'obj_key': 'GBARBOSA', 'grupo': 'normal', 'icon': 'G', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'ASSAI': {'nome': 'Assaí', 'obj_key': 'ASSAI', 'grupo': 'normal', 'icon': 'A', 'cor': '#C0200E', 'bg': '#FEF0EE'},
    'SAMS': {'nome': "Sam's Club", 'obj_key': 'SAMS', 'grupo': 'normal', 'icon': 'S', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'ATAKAREJO': {'nome': 'Atakarejo', 'obj_key': 'ATAKAREJO', 'grupo': 'normal', 'icon': 'AT', 'cor': '#E84B1A', 'bg': '#FEF3EE'},
    'AUTOZONE': {'nome': 'AutoZone', 'obj_key': None, 'grupo': 'normal', 'icon': 'AZ', 'cor': '#C0200E', 'bg': '#FEF0EE'},
    'OBRAMAX': {'nome': 'Obramax', 'obj_key': 'OBRAMAX', 'grupo': 'normal', 'icon': 'OB', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'TIM': {'nome': 'TIM', 'obj_key': None, 'grupo': 'normal', 'icon': 'T', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'GPA': {'nome': 'GPA', 'obj_key': 'GPA', 'grupo': 'normal', 'icon': 'GP', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'JLL': {'nome': 'JLL', 'obj_key': 'JLL', 'grupo': 'normal', 'icon': 'J', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'TENDA': {'nome': 'Tenda', 'obj_key': 'TENDA', 'grupo': 'normal', 'icon': 'TE', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'MERCANTIL': {'nome': 'Mercantil (Mercado)', 'obj_key': 'MERCANTIL', 'grupo': 'normal', 'icon': 'ME', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'BANCO MERCANTIL': {'nome': 'Banco Mercantil', 'obj_key': 'BANCO MERCANTIL', 'grupo': 'normal', 'icon': 'BM', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'ACCENTURE': {'nome': 'Accenture', 'obj_key': 'ACCENTURE', 'grupo': 'normal', 'icon': 'AC', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'DIA': {'nome': 'DIA', 'obj_key': None, 'grupo': 'normal', 'icon': 'D', 'cor': '#E84B1A', 'bg': '#FEF3EE'},
    'BAUDUCCO': {'nome': 'Bauducco', 'obj_key': None, 'grupo': 'normal', 'icon': 'BA', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'SIEMENS': {'nome': 'Siemens', 'obj_key': 'SIEMENS', 'grupo': 'normal', 'icon': 'SI', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'SINDILOJAS': {'nome': 'Sindilojas', 'obj_key': 'SINDILOJAS', 'grupo': 'normal', 'icon': 'SN', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'SUMERBOL': {'nome': 'Sumerbol', 'obj_key': 'SUMERBOL', 'grupo': 'normal', 'icon': 'SU', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'SAPATARIA': {'nome': 'Sapataria Nova', 'obj_key': 'SAPATARIA', 'grupo': 'normal', 'icon': 'SA', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'PORTO': {'nome': 'Porto Seguro', 'obj_key': 'PORTO', 'grupo': 'normal', 'icon': 'PS', 'cor': '#C0200E', 'bg': '#FEF0EE'},
    'ROVERI': {'nome': 'Roveri', 'obj_key': 'ROVERI', 'grupo': 'normal', 'icon': 'RO', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'GIGA': {'nome': 'Cencosud Giga', 'obj_key': 'GIGA', 'grupo': 'normal', 'icon': 'GI', 'cor': '#E84B1A', 'bg': '#FEF3EE'},
    'COBASI': {'nome': 'Cobasi', 'obj_key': None, 'grupo': 'normal', 'icon': 'CO', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'BIG': {'nome': 'BIG', 'obj_key': None, 'grupo': 'normal', 'icon': 'BI', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'SMS': {'nome': 'SMS', 'obj_key': None, 'grupo': 'normal', 'icon': 'SM', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'IMOPAR': {'nome': 'Imopar (Shopping Paseo)', 'obj_key': None, 'grupo': 'normal', 'icon': 'IM', 'cor': '#8A96B0', 'bg': '#F0F2F7'},
    'OUTLET': {'nome': 'Outlet Premium Itaquaquecetuba', 'obj_key': None, 'grupo': 'normal', 'icon': 'OU', 'cor': '#8A96B0', 'bg': '#F0F2F7'},
    'VIVARA': {'nome': 'Vivara', 'obj_key': None, 'grupo': 'normal', 'icon': 'VI', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'LG': {'nome': 'LG', 'obj_key': None, 'grupo': 'normal', 'icon': 'LG', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'GM': {'nome': 'GM', 'obj_key': None, 'grupo': 'normal', 'icon': 'GM', 'cor': '#1B3A6B', 'bg': '#EEF2FA'},
    'OKEAN': {'nome': 'Okean', 'obj_key': None, 'grupo': 'normal', 'icon': 'OK', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'RD SAUDE': {'nome': 'RD Saúde', 'obj_key': None, 'grupo': 'normal', 'icon': 'RS', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    # clientes novos, encontrados na leitura da planilha real em 2026-08-11:
    'BRETAS': {'nome': 'Bretas', 'obj_key': None, 'grupo': 'normal', 'icon': 'BR', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
    'COGNA': {'nome': 'Cogna', 'obj_key': None, 'grupo': 'normal', 'icon': 'CG', 'cor': '#5B35B0', 'bg': '#F0ECFC'},
    'ENGEMON': {'nome': 'Engemon', 'obj_key': None, 'grupo': 'normal', 'icon': 'EG', 'cor': '#0A7B8A', 'bg': '#E6F6F8'},
    'RD': {'nome': 'RD (Raia/Drogasil)', 'obj_key': 'RD', 'grupo': 'normal', 'icon': 'RD', 'cor': '#E84B1A', 'bg': '#FEF3EE'},
    'SENAC': {'nome': 'Senac', 'obj_key': 'SENAC', 'grupo': 'normal', 'icon': 'SC', 'cor': '#B07000', 'bg': '#FFF8E6'},
    'HAPVIDA': {'nome': 'Hapvida', 'obj_key': None, 'grupo': 'normal', 'icon': 'HV', 'cor': '#0E8A5A', 'bg': '#EAF7F1'},
}

# Planilha de OSP: cada aba = 1 cliente, coluna "OSP" = codigo unico da unidade
# (mesmo valor que aparece em "OS MANENG" na planilha de faturamento).
# name_col = indice (0-based) da coluna com o nome/endereco da unidade nessa aba.
OSP_SHEET_CONFIG = {
    'SAMS': {'sigla': 'SAMS', 'name_col': 2},
    'CARREFOUR ': {'sigla': 'CRF', 'name_col': 1},
    'ASSAI': {'sigla': 'ASSAI', 'name_col': 1},
    'ATAKAREJO': {'sigla': 'ATAKAREJO', 'name_col': 1},
    'AUTOZONE': {'sigla': 'AUTOZONE', 'name_col': 1},
    'BANCO MERCANTIL': {'sigla': 'BANCO MERCANTIL', 'name_col': 1},
    'BOTICARIO': {'sigla': 'BOT', 'name_col': 4},
    'BRADESCO - AGENCIAS': {'sigla': 'AG BRAD', 'name_col': 1},
    'BRADESCO - RESERVA TECNICA': {'sigla': 'BRAD', 'name_col': 1},
    'BRADESCO - IN HAUS': {'sigla': 'IN HAUS', 'name_col': 1},
    'BRADESCO - FUNDAÇÃO': {'sigla': 'FUND BRAD', 'name_col': 1},
    'CENCOSUD - GIGA': {'sigla': 'GIGA', 'name_col': 1},
    'CENCOSUD - MERCANTIL': {'sigla': 'MERCANTIL', 'name_col': 1},
    'DROGASIL': {'sigla': 'RD', 'name_col': 1},
    'CENCOSUD - GBARBOSA': {'sigla': 'GBARBOSA', 'name_col': 1},
    'DIA': {'sigla': 'DIA', 'name_col': 1},
    'GPA': {'sigla': 'GPA', 'name_col': 1},
    'HAPVIDA': {'sigla': 'HAPVIDA', 'name_col': 1},
    'JLL': {'sigla': 'JLL', 'name_col': 1},
    'LEROY MERLIN': {'sigla': 'LM', 'name_col': 1, 'servico_col': 3, 'residente_sigla': 'LM_RESIDENTES'},
    'OBRAMAX': {'sigla': 'OBRAMAX', 'name_col': 1},
    'SAPATARIA NOVA': {'sigla': 'SAPATARIA', 'name_col': 2},
    'SENAC': {'sigla': 'SENAC', 'name_col': 1},
    'SMARTFIT': {'sigla': 'SMTF', 'name_col': 1},
    'TENDA ': {'sigla': 'TENDA', 'name_col': 1},
    'ZAMP': {'sigla': 'ZAMP', 'name_col': 1},
}

# aba "DIVERSOS": mistura varios clientes pequenos na mesma aba -> mapeamento manual por OSP
DIVERSOS_OSP_TO_SIGLA = {
    '337': 'SIEMENS', '823': 'CRF', '1804': 'IMOPAR', '1806': 'ACCENTURE',
    '1827': 'BAUDUCCO', '1840': 'OUTLET',
    '1820/1': 'SUMERBOL', '1820/2': 'SUMERBOL', '1820/3': 'SUMERBOL',
    '1823/1': 'SINDILOJAS', '1823/2': 'SINDILOJAS', '1823/3': 'SINDILOJAS',
    '1824/1': 'ROVERI', '1824/2': 'ROVERI',
    '1828/1': 'TIM', '1828/2': 'TIM', '1828/3': 'TIM', '1828/4': 'TIM',
    '1829/1': 'PORTO', '1829/2': 'PORTO', '1830/1': 'PORTO', '1830/2': 'PORTO',
    '653 / 654': 'CRF',
    '1842': 'COGNA',
}

# tradução dos rótulos do bloco "Modelo / Metas Valores" da aba Objetivo -> chaves usadas no dashboard
GRUPO_KEY_MAP = {
    'PREVENTIVA TOTAL': 'PREVENTIVA TOTAL',
    'CORRETIVA': 'CORRETIVA',
    'ENGENHARIA': 'ENGENHARIA',
    'ZAMP - PREVENTIVA': 'ZAMP_P',
    'ZAMP - CORRETIVA': 'ZAMP_C',
    'BRADESCO AGENCIAS': 'BRADESCO AGENCIAS',
}


def log(msg):
    print(msg, file=sys.stderr)


def download(url, path):
    log(f"Baixando {url} ...")
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    if not data.startswith(b'PK'):
        raise RuntimeError(f"Download de {url} não retornou um .xlsx válido (sem login/permissão?)")
    with open(path, 'wb') as f:
        f.write(data)
    log(f"OK, {len(data)} bytes salvos em {path}")


def normalize_osp(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    s = re.sub(r'^OSP\s*', '', s, flags=re.IGNORECASE)  # ex.: "OSP 1832/03" -> "1832/03"
    s = re.sub(r'[A-Za-z]+$', '', s).strip()  # ex.: "1832/08A" -> "1832/08" (variacao de 2a via de nota)
    if '/' in s:
        partes = [p.lstrip('0') or '0' if p.isdigit() else p for p in s.split('/')]
        s = '/'.join(partes)  # ex.: "1794/064" -> "1794/64" (zero a esquerda inconsistente entre cadastro e nota)
    elif s.isdigit():
        s = str(int(s))
    return s


def parse_osp_registry(warnings):
    import openpyxl
    wb = openpyxl.load_workbook(OSP_PATH, data_only=True, read_only=True)
    registro = {}  # sigla -> [[osp, nome], ...]

    def add(sigla, osp, nome):
        if not osp:
            return
        registro.setdefault(sigla, []).append([osp, (nome or '').strip()])

    for sheet_name, cfg in OSP_SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            warnings.append(f"OSP: aba '{sheet_name}' não encontrada na planilha")
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        header_idx = next((i for i, r in enumerate(rows[:5]) if r and r[0] and str(r[0]).strip().upper() == 'OSP'), None)
        if header_idx is None:
            warnings.append(f"OSP: cabeçalho 'OSP' não encontrado na aba '{sheet_name}'")
            continue
        for r in rows[header_idx + 1:]:
            if not r or r[0] in (None, ''):
                continue
            if str(r[0]).strip().upper() == 'OSP':
                continue  # cabecalho repetido no meio da aba
            osp = normalize_osp(r[0])
            nome = r[cfg['name_col']] if len(r) > cfg['name_col'] else None
            sigla = cfg['sigla']
            if 'servico_col' in cfg and len(r) > cfg['servico_col']:
                servico = str(r[cfg['servico_col']] or '').strip().upper()
                if 'RESIDENTE' in servico:
                    sigla = cfg['residente_sigla']
            add(sigla, osp, nome)

    ws = wb['DIVERSOS']
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    header_idx = next((i for i, r in enumerate(rows[:5]) if r and r[0] and str(r[0]).strip().upper() == 'OSP'), None)
    for r in rows[header_idx + 1:]:
        if not r or r[0] in (None, ''):
            continue
        osp = normalize_osp(r[0])
        sigla = DIVERSOS_OSP_TO_SIGLA.get(osp)
        if sigla is None:
            warnings.append(f"OSP (DIVERSOS): código '{osp}' ({r[1]!r}) sem mapeamento de SIGLA — ignorado")
            continue
        add(sigla, osp, r[1])

    return registro


def parse_valor(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).strip()
    s = re.sub(r'[Rr]\$', '', s).strip()
    if re.search(r',\d{2}$', s) and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif re.search(r',\d{2}$', s):
        s = s.replace(',', '.')
    else:
        s = s.replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def build():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    meses_disponiveis = [m for m in MESES_ORDEM if m in wb.sheetnames]
    log(f"Abas de mês encontradas: {meses_disponiveis}")

    agregado = {}
    faturado_osp = {}  # mes -> sigla -> [osp, ...] (unidades com preventiva/P-ENG faturada)
    warnings = []
    siglas_nao_mapeadas = set()

    # le a Planilha de OSP primeiro: usamos o cadastro de unidades (fonte confiavel) pra
    # separar Leroy Merlin normal de Residentes, em vez de depender do texto de OBSERVACAO
    # (que nem sempre vem preenchido do jeito esperado).
    registro_osp = parse_osp_registry(warnings)
    lm_residente_codes = {osp for osp, _nome in registro_osp.get('LM_RESIDENTES', [])}

    for mes in meses_disponiveis:
        ws = wb[mes]
        agregado[mes] = {}
        faturado_osp[mes] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or len(row) < 9:
                continue
            valor, sigla, serv, os_maneng = row[2], row[6], row[7], row[8]
            obs = row[5] if len(row) > 5 else None
            if serv is None or sigla is None:
                continue
            serv = str(serv).strip().upper()
            if serv not in ('P', 'C', 'E', 'P-ENG'):
                continue
            sigla = str(sigla).strip()
            sigla = ALIAS.get(sigla, sigla)
            if sigla in SIGLAS_IGNORADAS:
                continue
            if sigla == 'LM' and normalize_osp(os_maneng) in lm_residente_codes:
                sigla = 'LM_RESIDENTES'
            cat = 'E' if serv == 'P-ENG' else serv
            val = parse_valor(valor)
            if val is None:
                warnings.append(f"{mes}: VALOR não numérico em linha SIGLA={sigla} SERV={serv}: {valor!r}")
                continue
            d = agregado[mes].setdefault(sigla, {'P': 0.0, 'C': 0.0, 'E': 0.0})
            d[cat] = round(d[cat] + val, 2)
            if sigla not in SIGLA_MAP:
                siglas_nao_mapeadas.add(sigla)
            if serv in ('P', 'P-ENG'):
                osp_norm = normalize_osp(os_maneng)
                if osp_norm:
                    faturado_osp[mes].setdefault(sigla, set()).add(osp_norm)

    # aba Objetivo: metas por cliente + metas por grupo
    ws = wb['Objetivo']
    obj_cliente = {}
    obj_grupo = {}
    modo_grupo = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        chave = str(row[0]).strip()
        if chave.lower() == 'cliente' or chave.lower() == 'modelo':
            modo_grupo = (chave.lower() == 'modelo')
            continue
        if chave.lower() == 'total geral':
            continue
        valor = row[1]
        if valor is None:
            continue
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            continue
        if modo_grupo:
            gk = GRUPO_KEY_MAP.get(chave)
            if gk is None:
                warnings.append(f"Objetivo (grupo): rótulo não reconhecido '{chave}' ignorado")
                continue
            obj_grupo[gk] = valor
        else:
            if chave == 'SHOP. PAMPLONA':
                chave = 'CARREFOUR'
                valor += obj_cliente.get('CARREFOUR', 0.0)
            obj_cliente[chave] = valor

    for k in ('PREVENTIVA TOTAL', 'CORRETIVA', 'ENGENHARIA', 'ZAMP_P', 'ZAMP_C', 'BRADESCO AGENCIAS'):
        if k not in obj_grupo:
            warnings.append(f"Meta de grupo '{k}' não encontrada na aba Objetivo")

    # cliente novo na planilha, sem entrada manual em SIGLA_MAP -> cadastra automaticamente
    # usando o proprio nome (nunca cai num balaio generico de "Outros").
    PALETA_AUTO = [
        ('#1B3A6B', '#EEF2FA'), ('#0E8A5A', '#EAF7F1'), ('#E84B1A', '#FEF3EE'),
        ('#5B35B0', '#F0ECFC'), ('#0A7B8A', '#E6F6F8'), ('#B07000', '#FFF8E6'),
        ('#C0200E', '#FEF0EE'),
    ]
    for i, sigla in enumerate(sorted(siglas_nao_mapeadas)):
        cor, bg = PALETA_AUTO[i % len(PALETA_AUTO)]
        nome = sigla if len(sigla) <= 3 else sigla.title()
        SIGLA_MAP[sigla] = {'nome': nome, 'obj_key': None, 'grupo': 'normal', 'icon': sigla[:2].upper(), 'cor': cor, 'bg': bg}
        log(f"Cliente novo detectado e cadastrado automaticamente: SIGLA='{sigla}' -> nome '{nome}' (ainda sem meta)")

    # checagem: toda meta de cliente deveria estar referenciada por algum SIGLA_MAP.obj_key
    obj_keys_usadas = {v['obj_key'] for v in SIGLA_MAP.values() if v['obj_key']}
    metas_orfas = sorted(set(obj_cliente.keys()) - obj_keys_usadas)

    if metas_orfas:
        warnings.append(f"Metas na aba Objetivo sem nenhum cliente do painel apontando para elas: {metas_orfas}")

    # mês padrão = último mês (na ordem) que tem faturamento real
    mes_atual = meses_disponiveis[0]
    for m in meses_disponiveis:
        total_mes = sum(sum(v.values()) for v in agregado[m].values())
        if total_mes > 0:
            mes_atual = m

    # (registro_osp ja foi lido no inicio da funcao, usado tambem para achar pendencias de faturamento)
    siglas_osp_nao_mapeadas = sorted(set(registro_osp.keys()) - set(SIGLA_MAP.keys()))
    if siglas_osp_nao_mapeadas:
        warnings.append(f"OSP: SIGLAs com unidades cadastradas mas SEM entrada em SIGLA_MAP: {siglas_osp_nao_mapeadas}")

    faturado_osp_json = {m: {s: sorted(v) for s, v in d.items()} for m, d in faturado_osp.items()}

    tz = timezone(timedelta(hours=-3))
    agora = datetime.now(tz).strftime('%d/%m/%Y %H:%M')

    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        html = f.read()

    def inject(marker, value_json):
        nonlocal html
        if marker not in html:
            raise RuntimeError(f"Marcador {marker} não encontrado no template")
        html = html.replace(marker, value_json, 1)

    inject('/*ALIAS_JSON*/', json.dumps(ALIAS, ensure_ascii=False))
    inject('/*MESES_JSON*/', json.dumps(meses_disponiveis, ensure_ascii=False))
    inject('/*MES_ATUAL_JSON*/', json.dumps(mes_atual, ensure_ascii=False))
    inject('/*AGREGADO_JSON*/', json.dumps(agregado, ensure_ascii=False))
    inject('/*OBJ_CLIENTE_JSON*/', json.dumps(obj_cliente, ensure_ascii=False))
    inject('/*OBJ_GRUPO_JSON*/', json.dumps(obj_grupo, ensure_ascii=False))
    inject('/*SIGLA_MAP_JSON*/', json.dumps(SIGLA_MAP, ensure_ascii=False))
    inject('/*REGISTRO_OSP_JSON*/', json.dumps(registro_osp, ensure_ascii=False))
    inject('/*FATURADO_OSP_JSON*/', json.dumps(faturado_osp_json, ensure_ascii=False))
    html = html.replace('/*GENERATED_AT*/', f'gerado em {agora} (America/Sao_Paulo)')
    html = html.replace('/*ATUALIZADO_EM*/', agora)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)

    log(f"\nHTML final escrito em {OUTPUT_PATH}")
    log(f"Mês padrão selecionado: {mes_atual}")
    log("\nTotal faturado por mês (P+C+E, todas as siglas):")
    for m in meses_disponiveis:
        tot = sum(sum(v.values()) for v in agregado[m].values())
        log(f"  {m}: R$ {tot:,.2f}")

    if warnings:
        log("\n=== AVISOS (revisar) ===")
        for w in warnings:
            log(f"  - {w}")
    else:
        log("\nNenhum aviso — todas as SIGLAs e metas foram mapeadas corretamente.")

    return {
        'mes_atual': mes_atual,
        'meses': meses_disponiveis,
        'warnings': warnings,
        'output_path': OUTPUT_PATH,
    }


if __name__ == '__main__':
    download(XLSX_URL, XLSX_PATH)
    download(OSP_URL, OSP_PATH)
    build()
